#UTILS FILE
#created: 09/09/2021

#imports
import numpy as np
import SimpleITK as sitk
from skimage.transform import rescale
from scipy.ndimage.measurements import center_of_mass
import torch
from functools import reduce
from operator import mul
import cv2
import sklearn
from decimal import Decimal
import matplotlib.pyplot as plt
from torchvision import models
from scipy import ndimage
import albumentations as A
from albumentations.pytorch import ToTensor
import os
from openpyxl import load_workbook
import pandas as pd
import math
from model import neckNavigator

#constants
window = 350
level = 50

#general utils functions
def GetTargetCoords(target):
  coords = center_of_mass(target) 
  return coords #z,x,y

def GetSliceNumber(segment):
  slice_number = []
  weights = []
  max_range = len(segment)
  for x in range(0,max_range):
    seg_slice = segment[x,...]
    val = np.sum(seg_slice)
    if val != 0:
      slice_number.append(x)
      weights.append(val)
  slice_no = do_it_urself_round(np.average(slice_number, weights = weights), decimals=3)
  return slice_no

def do_it_urself_round(number, decimals = 0, is_array = False):
  float_num = number
  dec_num =  Decimal(str(number))
  Ddecimal = Decimal(str(decimals))
  round_num = round(dec_num, decimals)
  diff = np.abs(dec_num - round_num)
  round_diff = Decimal(str(0.5/(10**decimals)))
  if (diff == round_diff and dec_num >= round_num):
    round_num += (1/(10**(Ddecimal))) 
  if decimals == 0:
    return int(round_num)
  else:
    return float(round_num)

def flat_softmax(inp):
    """Compute the softmax with all but the first two tensor dimensions combined."""
    orig_size = inp.size()
    flat = inp.view(-1, reduce(mul, orig_size[2:]))
    flat = torch.nn.functional.softmax(flat, -1)
    return flat.view(*orig_size)

###*** LOAD ***###
def load(path):    
    # Load input and target
    x = sitk.ReadImage(path, imageIO="NiftiImageIO")
    #saving the spacing
    voxel_dim = [(x.GetSpacing())[0],(x.GetSpacing())[1],(x.GetSpacing())[2]]
    
    input_data = {'input': x, 'voxel_dims': np.array(voxel_dim).astype(np.float32)}  
    return input_data

def get_patient_id(dir):
  #_,name = os.path.split()
  patients_path = [os.path.join(dir, x) for x in os.listdir(dir)]
  names = [os.path.splitext(y)[0] for y in os.listdir(dir)]
  names = [os.path.splitext(y)[0] for y in names]
  return patients_path[237:], names[237:]

###*** PRE-PROCESSING 1 ***###
def flip(im):
    im = np.flip(im, axis=0)        # flip CC
    im = np.flip(im, axis=2)        # flip LR
    return im

def window_level(inp: np.array):
    vmax = level/2 + window
    vmin = level/2-window
    thresh = 1000
    inp[inp > thresh] = 0
    inp[inp > vmax] = vmax
    inp[inp < vmin] = vmin
    return inp

def normalize_01(inp: np.ndarray):
    """Squash image input to the value range [0, 1] (plus clipping)"""
    inp = (np.nan_to_num(inp)).astype(np.float64)
    shape = inp.shape
    inp_new = np.round(sklearn.preprocessing.minmax_scale(inp.ravel(), feature_range=(0,1)), decimals = 10).reshape(shape)
    
    inp_out = inp_new
    return inp_out

def cropping(inp, bone_mask = None, threeD = True):
    #working one but z axis crop needs improving
    x = inp
    _,threshold = cv2.threshold(x,200,0,cv2.THRESH_TOZERO)
    coords = center_of_mass(threshold)
    print(coords)
    if threeD: x_ind, y_ind = 1, 2
    else: x_ind, y_ind = 0, 1
    if threeD: size =126
    else: size = 130 #110 would be nice
    x_min = do_it_urself_round(((coords[x_ind] - size)+126)/2) #126
    x_max = do_it_urself_round(((coords[x_ind] + size)+386)/2)
    y_min = do_it_urself_round(((coords[y_ind] - size)+126)/2)
    y_max = do_it_urself_round(((coords[y_ind] + size)+386)/2)
    
    if threeD == True:
      #z crop
      z_size = 112
      z_coords = {"z_min":x.shape[0]-z_size,"z_max":x.shape[0]}
      org_inp_size = x.shape
          
      if (z_size < x.shape[0]):
          #print("bigger", x.shape[0])
          if(x.shape[0] > do_it_urself_round(coords[0])+z_size): #190>87+112=199
              print("crop")
              z_coords = {"z_min": do_it_urself_round(coords[0]), "z_max": do_it_urself_round(coords[0])+z_size}

      else:
          print("too small: ", x.shape[0]) #107 112-107/2
          padded_arr = np.pad(x,((math.floor((z_size-x.shape[0])/2), math.ceil((z_size-x.shape[0])/2)), (0,0),(0,0)),'mean')
          
          inp = padded_arr
          z_coords = {"z_min": 0, "z_max": inp.shape[0]}
      
      x = inp[z_coords["z_min"]:z_coords["z_max"],x_min:x_max,y_min:y_max]
      cropped_info = [z_coords["z_min"], z_coords["z_max"], org_inp_size[0], x_min, x_max, org_inp_size[1], y_min, y_max, org_inp_size[2]]
      print(x.shape)
      return x, np.array(cropped_info)

    else:
      x = inp[x_min:x_max,y_min:y_max]
      if bone_mask is not None:
        cropped_bone = bone_mask[x_min:x_max,y_min:y_max]
        return x, cropped_bone
      else:
        return x

def preprocessing_1(image):
    x = image
    # check if flip required
    need_flip = False
    if x.GetDirection()[-1] == -1:
        need_flip = True
    x= sitk.GetArrayFromImage(x).astype(float)
    print("worldmatch info here:",x.shape, np.max(x), np.min(x))
    x-=1024 #worldmatch trickery

    # Preprocessing
    if need_flip == True:
        x= flip(x)

    x = window_level(x)
    x, crop_info = cropping(x)
    x = normalize_01(x)
    #downsampling to size -> [128,128,128]
    x = rescale(x, scale=((16/14),0.5,0.5), order=0, multichannel=False,  anti_aliasing=False)
    #save transforms
    transform_list = np.array([need_flip, crop_info])
    processed_data = {'input': x, 'transform': transform_list}  
    return processed_data

###*** NECK NAVIGATOR MODEL ***###
#MODEL SETUP
def setup_model(model, checkpoint_dir, device, load_prev = False, load_best = False, eval_mode = False):
  model.to(device)
  if load_prev == True:
    model.load_previous(checkpoint_dir, logger=None)
  if load_best == True:
    model.load_best(checkpoint_dir, logger = None)
  for param in model.parameters():
      param.requires_grad = True
  if eval_mode:
    model.eval()
  return model

def NeckNavigatorRun(model_dir, ct, device): #put one image through
    model = setup_model(neckNavigator(), model_dir, device, load_best = True, eval_mode=True)
    input_image = torch.from_numpy(ct).type(torch.FloatTensor)
    # creating channel dimension            
    input_image = input_image.unsqueeze(0).unsqueeze(0)
    input_image = input_image.type(torch.FloatTensor).to(device)
    print("NN input: ",input_image.shape)
    output = model(input_image)
    output = flat_softmax(output)
    test_output = output.squeeze().cpu().detach().numpy()
    pred = test_output.astype(np.float)
    return np.array(pred)

###*** POST-PROCESSING 1 ***###
def mrofsnart(msk, transforms, shape = 128):#transforms backwards
    #undo scale
    coords = GetTargetCoords(msk)
    #print(coords)
    mask = rescale(msk, scale=((14/16),2,2), order=0, multichannel=False,  anti_aliasing=False)
    coords = GetTargetCoords(mask)
    #undo crop
    #eg z crop [46,1] z=12  [[true, crop array] <- crop[zmin, zmax, org_size, xmin,...],]
    z = coords[0] + transforms[1][0]
    x = coords[1]
    y = coords[2]
    #print("Undo Scale: ", coords)
    x += transforms[1][3]
    y += transforms[1][6]
    #print("Undo Crop:", z, x, y)
    #undo flip if necessary
    if (transforms[0]==True):
        y_shape = transforms[1][8] -1
        y = y_shape - y

    if (transforms[0]==True):
        z_shape = transforms[1][2] -1 
        z = z_shape - z
    #print("Final Coords: ", z,x,y)
    return do_it_urself_round(x,1),do_it_urself_round(y,1),do_it_urself_round(z)

###*** SAVE NECK NAVIGATOR OUTPUT ***###
#slice number and sagital image and patient id
def arrange(input, ax, proj_order):
    #to return the projection in whatever order.
    av = np.average(input, axis = ax)
    mx = np.max(input, axis = ax)
    std = np.std(input, axis=ax)
    ord_list = [mx,av,std]
    ord_list[:] = [ord_list[i] for i in proj_order]
    out = np.stack((ord_list), axis=2)
    return out

def base_projections(inp, msk):
  axi,cor,sag = 0,1,2
  proj_order = [2,1,0]
  axial = arrange(inp, axi, proj_order)
  coronal = arrange(inp, cor, proj_order)
  sagital = arrange(inp, sag, proj_order)
  ax_mask = np.max(msk, axis = axi)
  cor_mask = np.max(msk, axis = cor)
  sag_mask = np.max(msk, axis = sag)
  #flip right way up
  coronal = coronal[::-1]
  sagital = sagital[::-1]
  cor_mask = cor_mask[::-1]
  sag_mask = sag_mask[::-1]
  image = (axial,coronal,sagital)
  mask = (ax_mask,cor_mask, sag_mask)
  return image, mask

def display_net_test(inps, msks,ax, shape = 128, z = None):
    coords_pred = GetTargetCoords(msks)  
    #make the figure
    # fig = plt.figure()
    image, pred = base_projections(inps, msks)
    slice_pred = shape - do_it_urself_round(GetSliceNumber(msks)) #as projections have different origin
    ax.set_title('C3 Location')
    ax.imshow(image[2])
    #plt.imshow(pred[2], cmap = 'cool', alpha=0.5) #maybe take this out
    ax.axhline(slice_pred, linewidth=2, c='y')
    if z is None: slice_no = slice_pred
    else: slice_no = do_it_urself_round(z)
    ax.text(0, slice_pred-5, "C3: "+ str(slice_no), color='w')
    ax.scatter(coords_pred[1], (128 - coords_pred[0]), c = 'r', s=20) #y,z
    ax.axis('off')
  

###*** PRE-PROCESSING 2 ***###
def window_level_norm(inp: np.array):
  vmax = level/2 + window
  vmin = level/2-window
  # thresh = 1000
  # inp[inp > thresh] = 0
  inp[inp > vmax] = vmax
  inp[inp < vmin] = vmin
  shape = inp.shape
  image_scaled = np.round(sklearn.preprocessing.minmax_scale(inp.ravel(), feature_range=(0,1)), decimals = 10).reshape(shape)
  return image_scaled

def extract_bone_masks(dcm_array, slice_number, threshold=200, radius=2, worldmatch=False):
    """
    Calculate 3D bone mask and remove from prediction. 
    Used to account for partial volume effect.
​
    Args:
        dcm_array - 3D volume
        slice number - slice where segmentation being performed
        threshold (in HU) - threshold above which values are set to 1 
        radius (in mm) - Kernel radius for binary dilation
    """
    #img = sitk.GetImageFromArray(dcm_array)
    img = dcm_array
    #crop_by = 5 #these two lines were meant to make it quicker.
    #img = img[...,(slice_number-crop_by):(slice_number+crop_by)]
    # Worldmatch tax
    if worldmatch:
        img -= 1024
    # Apply threshold
    bin_filt = sitk.BinaryThresholdImageFilter()
    bin_filt.SetOutsideValue(1)
    bin_filt.SetInsideValue(0)
    bin_filt.SetLowerThreshold(-1024)
    bin_filt.SetUpperThreshold(threshold)
    bone_mask = bin_filt.Execute(img)
    pix = bone_mask.GetSpacing()
    # Convert to pixels
    pix_rad = [int(radius//elem) for elem in pix]
    # Dilate mask
    dil = sitk.BinaryDilateImageFilter()
    dil.SetKernelType(sitk.sitkBall)
    dil.SetKernelRadius(pix_rad)
    dil.SetForegroundValue(1)
    dilated_mask = dil.Execute(bone_mask)
    im_arr = sitk.GetArrayFromImage(dilated_mask)[slice_number]
    return np.logical_not(im_arr)

def three_channel(ct_slice):
  slices_3chan = np.repeat(ct_slice[...,np.newaxis], 3, axis=-1)
  #bone_3chan = np.repeat(bone_mask[...,np.newaxis], 3, axis=-1)
  # apply filters to two channels
  slices_3chan[...,1] = ndimage.gaussian_laplace(slices_3chan[...,1], sigma=1)
  slices_3chan[...,2] = ndimage.sobel(slices_3chan[...,2])
  return slices_3chan.astype(np.float32) #, bone_3chan.astype(np.float32)

transform = A.Compose([A.Resize(260, 260), A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)), ToTensor()])
#240 crop would be nice but have to do some train tweaks think
def preprocessing_2(slice_no, ct, is_worldmatch = True):
  #extract slice
  ct_slice = sitk.GetArrayFromImage(ct)[slice_no]
  print("ct_info: ", ct_slice.shape, np.max(ct_slice),np.min(ct_slice))
  if is_worldmatch: ct_slice -=1024
  ct_slice = ct_slice.astype(float)
  #extract bone mask
  bone_mask = extract_bone_masks(ct, slice_no, worldmatch=is_worldmatch)
  print("bone_info",bone_mask.shape, np.unique(bone_mask))
  cropped_slice, cropped_bone = cropping(ct_slice, bone_mask, threeD=False)
  #print("crop: ",cropped_slice.shape)
  # plt.imshow(cropped_slice)
  # plt.imshow(cropped_bone.astype(float), alpha=0.5, cmap = "cool")
  # plt.show()
  wln_slice = window_level_norm(cropped_slice)
  preprocess2_info = {"slice": wln_slice, "bone":cropped_bone}
  return preprocess2_info

###*** MUSCLE MAPPER MODEL ***###
def set_up_MuscleMapper(model_path, device):
  #initilaise and load the model
  model = models.segmentation.fcn_resnet50(pretrained=False, num_classes=1)
  model.load_state_dict(torch.load(model_path, map_location=device))
  model.to(device)
  model.eval()
  return model

def MuscleMapperRun(ct_slice, model_path, device):
  #do three channels with filters # should this be before scaling
  three_chan = three_channel(ct_slice)
  transformed = transform(False, image = three_chan)
  transformed_im = transformed["image"]
  input_slice = transformed_im.unsqueeze(0).to(device)
  input_slice = input_slice.type(torch.float32)
  print("model input:", input_slice.shape)
  #put through model
  model = set_up_MuscleMapper(model_path, device)
  output = model(input_slice)["out"] 
  MM_ouput = output.detach().cpu().squeeze()
  # sigmoid and thresholding
  sigmoid = 1/(1 + np.exp(-MM_ouput))
  segment = (sigmoid > 0.5).float().numpy()
  return np.array(segment).astype(int)

###*** POST-PROCESSING 2 ***###
def getDensity(image, mask, area, label=1):
  if image.shape != (len(image),1,...):
    mask = np.squeeze(mask)
  return float(np.mean(image[np.where(mask == 1)]))

def getArea(image, mask, area, label=1, thresholds = None):
  sMasks = (mask == label)
  threshold = np.logical_and(image > (thresholds[0]), image <  (thresholds[1]))
  tmask = np.logical_and(sMasks, threshold)
  return np.count_nonzero(tmask) * area

def postprocessing_2(org_ct, slice_no, pred, bone_mask, dims, is_worldmatch = False):
  ct_slice = (sitk.GetArrayFromImage(org_ct)[slice_no]).astype(float)
  print("postProcessing: ", ct_slice.shape, np.max(ct_slice), np.min(ct_slice))
  if is_worldmatch: ct_slice -= 1024
  cropped_slice = cropping(ct_slice, threeD=False)
  #remove bone masks
  pred_slb = np.logical_and(pred, bone_mask)
  #calc SMA/SMI and SMD
  pixel_area = dims[0]*dims[1]*(0.1*0.1) #mm^2 -> cm^2
  SMA = getArea(cropped_slice, pred_slb, pixel_area, thresholds=(-30, +130))
  SMD = getDensity(cropped_slice, pred_slb, pixel_area)
  return do_it_urself_round(SMA,3), do_it_urself_round(SMD,3)

###*** SAVE MUSCLE MAPPER OUTPUT ***###
#segment and patient ID and SMA/SMI and SMD
def display_slice(ct_slice, segment, ax):
  #fig = plt.figure()
  ax.imshow(ct_slice, cmap = plt.cm.gray)#, vmin = level/2 - window, vmax = level/2 + window
  seg = segment.astype(float)
  seg[seg == 0.0] = np.nan
  ax.imshow(seg, cmap = plt.cm.autumn, alpha = 0.5)
  ax.set_title("Skeletal Muscle")
  #plt.savefig(save_loc)
  ax.axis('off')

def save_figs(ct, ct_slice, segment, slice_pred, save_loc, id):
  fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(8, 4))
  fig.suptitle('Patient: '+ id)
  display_net_test(ct, slice_pred, ax1)
  display_slice(ct_slice, segment, ax2)
  #plt.show()
  plt.savefig(save_loc + f'image_{id}.png')
  plt.close()
  return


#saving to excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                      truncate_sheet=False, 
                      **to_excel_kwargs):
  """
  Append a DataFrame [df] to existing Excel file [filename]
  into [sheet_name] Sheet.
  If [filename] doesn't exist, then this function will create it.

  @param filename: File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
  @param df: DataFrame to save to workbook
  @param sheet_name: Name of sheet which will contain DataFrame.
                      (default: 'Sheet1')
  @param startrow: upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
  @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                          before writing DataFrame to Excel file
  @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                          [can be a dictionary]
  @return: None

  Usage examples:

  >>> append_df_to_excel('d:/temp/test.xlsx', df)

  >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

  >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                          index=False)

  >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                          index=False, startrow=25)

  (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
  """
  # Excel file doesn't exist - saving and exiting
  if not os.path.isfile(filename):
      df.to_excel(
          filename,
          sheet_name=sheet_name, 
          startrow=startrow if startrow is not None else 0, 
          **to_excel_kwargs)
      return
  
  # ignore [engine] parameter if it was passed
  if 'engine' in to_excel_kwargs:
      to_excel_kwargs.pop('engine')

  writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

  # try to open an existing workbook
  writer.book = load_workbook(filename)
  
  # get the last row in the existing Excel sheet
  # if it was not specified explicitly
  if startrow is None and sheet_name in writer.book.sheetnames:
      startrow = writer.book[sheet_name].max_row

  # truncate sheet
  if truncate_sheet and sheet_name in writer.book.sheetnames:
      # index of [sheet_name] sheet
      idx = writer.book.sheetnames.index(sheet_name)
      # remove [sheet_name]
      writer.book.remove(writer.book.worksheets[idx])
      # create an empty sheet [sheet_name] using old index
      writer.book.create_sheet(sheet_name, idx)
  
  # copy existing sheets
  writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

  if startrow is None:
      startrow = 0
  else:
    header = None
    to_excel_kwargs.update({'header':header})

  # write out the new sheet
  df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
  
  # save the workbook
  writer.save()